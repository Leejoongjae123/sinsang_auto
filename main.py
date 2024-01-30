import openpyxl
import pandas as pd
from pyautogui import size
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
import subprocess
import shutil
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities
from bs4 import BeautifulSoup
import time
import datetime
import pyautogui
import pyperclip
import csv
import sys
import os
import math
import requests
import re
import random
import chromedriver_autoinstaller
from PyQt5.QtWidgets import QWidget, QApplication, QTreeView, QFileSystemModel, QVBoxLayout, QPushButton, QInputDialog, \
    QLineEdit, QMainWindow, QMessageBox, QFileDialog
from PyQt5.QtCore import QCoreApplication
from selenium.webdriver import ActionChains
from datetime import datetime, date, timedelta
import numpy
import datetime
# from window import Ui_MainWindow
from PyQt5.QtWidgets import *
from PyQt5.QtCore import *
import json
import pprint
import os
import gspread
from oauth2client.service_account import ServiceAccountCredentials
import smtplib  # SMTP 사용을 위한 모듈
from email.mime.multipart import MIMEMultipart  # 메일의 Data 영역의 메시지를 만드는 모듈
from email.mime.text import MIMEText  # 메일의 본문 내용을 만드는 모듈
from email.mime.base import MIMEBase
from email import encoders


def ConvertToTwoDigits(num):
    # 숫자를 문자열로 변환
    num_str = str(num)
    # 1자리 숫자인 경우 앞에 0을 추가하여 2자리로 만듦
    if len(num_str) == 1:
        num_str = '0' + num_str
    return num_str

def is_time_between_8pm_and_5am():
    # 현재 시간을 가져옵니다
    current_time = datetime.datetime.now().time()

    # 저녁 8시와 새벽 5시를 정의합니다
    time_8_pm = datetime.datetime.strptime("20:00", "%H:%M").time()
    time_5_am = datetime.datetime.strptime("05:00", "%H:%M").time()

    # 현재 시간이 저녁 8시와 새벽 5시 사이인지 확인합니다
    if time_8_pm <= current_time or current_time <= time_5_am:
        return True
    else:
        return False

def GetToken(id, pw):
    headers = {
        'authority': 'abara.sinsang.market',
        'accept': 'application/json, text/plain, */*',
        'accept-language': 'ko-KR,ko;q=0.9,en-US;q=0.8,en;q=0.7',
        'content-type': 'application/json',
        'global': 'en',
        'origin': 'https://sinsangmarket.kr',
        'referer': 'https://sinsangmarket.kr/',
        'sec-ch-ua': '"Not.A/Brand";v="8", "Chromium";v="114", "Google Chrome";v="114"',
        'sec-ch-ua-mobile': '?0',
        'sec-ch-ua-platform': '"Windows"',
        'sec-fetch-dest': 'empty',
        'sec-fetch-mode': 'cors',
        'sec-fetch-site': 'cross-site',
        'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/114.0.0.0 Safari/537.36',
    }

    json_data = {
        'user': str(id),
        'password': str(pw),
    }

    response = requests.post('https://abara.sinsang.market/api/v1/session', headers=headers, json=json_data)
    response.raise_for_status()
    result = json.loads(response.text)
    accessToken = result['content']['accessToken']

    return accessToken

    # Note: json_data will not be serialized by requests
    # exactly as it was in the original request.
    # data = '{"user":"sjst","password":"mtqpxwuz"}'
    # response = requests.post('https://abara.sinsang.market/api/v1/session', headers=headers, data=data)


def GetProuctList(accessToken, shopId, size):
    headers = {
        'authority': 'ip-api.sinsang.market',
        'accept': 'application/json, text/plain, */*',
        'accept-language': 'ko-KR,ko;q=0.9,en-US;q=0.8,en;q=0.7',
        'access-token': str(accessToken),
        'global': 'en',
        'origin': 'https://sinsangmarket.kr',
        'platform': 'WEB',
        'referer': 'https://sinsangmarket.kr/',
        'sec-ch-ua': '"Not.A/Brand";v="8", "Chromium";v="114", "Google Chrome";v="114"',
        'sec-ch-ua-mobile': '?0',
        'sec-ch-ua-platform': '"Windows"',
        'sec-fetch-dest': 'empty',
        'sec-fetch-mode': 'cors',
        'sec-fetch-site': 'cross-site',
        'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/114.0.0.0 Safari/537.36',
    }

    size = int(size) + random.randint(-3, 3)
    params = {
        'page': '1',
        'size': str(size),
        'sort': 'DATE',
        'isPublic': 'true',
    }

    response = requests.get('https://ip-api.sinsang.market/search/v2/product/wholesale/{}'.format(shopId),
                            params=params, headers=headers)
    response.raise_for_status()
    print("statusCode:", response.status_code)
    result = json.loads(response.text)
    # pprint.pprint(result)
    storeNameWeb = result['content']['items'][0]['storeName']
    items = result['content']['items']
    dataList = []
    for item in items:
        # print(item)
        productId = item['wgIdx']
        dataList.append(productId)
    return dataList, storeNameWeb


def GetDetail(accessToken, productId, productCount, shopId, timeCreate):
    headers = {
        'authority': 'abara.sinsang.market',
        'accept': 'application/json, text/plain, */*',
        'accept-language': 'ko-KR,ko;q=0.9,en-US;q=0.8,en;q=0.7',
        'access-token': str(accessToken),
        'global': 'en',
        'if-none-match': 'W/"dfbd2af4b40778b9e6c925f09b790c3c"',
        'origin': 'https://sinsangmarket.kr',
        'platform': 'WEB',
        'referer': 'https://sinsangmarket.kr/',
        'sec-ch-ua': '"Not.A/Brand";v="8", "Chromium";v="114", "Google Chrome";v="114"',
        'sec-ch-ua-mobile': '?0',
        'sec-ch-ua-platform': '"Windows"',
        'sec-fetch-dest': 'empty',
        'sec-fetch-mode': 'cors',
        'sec-fetch-site': 'cross-site',
        'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/114.0.0.0 Safari/537.36',
        'uuid': '',
    }

    response = requests.get('https://abara.sinsang.market/api/v1/goods/{}/detail'.format(productId), headers=headers)
    response.raise_for_status()
    result = json.loads(response.text)['content']
    # pprint.pprint(result)
    name = result['name']
    print("name:", name)
    price = result['originPrice']
    print("price:", price)

    # 치환명 가져오기
    fname = 'Category.xlsx'
    changeList = CategoryChangeList(fname)

    cdIdx = result['cdIdx']
    category = GetCategory(cdIdx)
    # ----------------------------치환식------------------------
    for changeElem in changeList:
        category = category.replace(changeElem['target'], changeElem['changeName'])
    # ---------------------------------------------------------------
    print('category:', category)
    options = result['options']
    colorNameList = []
    sizeNameList = []
    colors = result['colorInfo']
    for color in colors:
        colorNameList.append(color['color'])
    colorNames = "|".join(colorNameList)

    # ----------------------------치환식------------------------
    for changeElem in changeList:
        colorNames = colorNames.replace(changeElem['target'], changeElem['changeName'])
    # ---------------------------------------------------------------
    print('colorNames:', colorNames)

    sizes = result['size']
    for size in sizes:
        sizeNameList.append(size)
    # print(sizeNameList)
    sizeNames = "|".join(sizeNameList)

    # ----------------------------치환식------------------------
    for changeElem in changeList:
        sizeNames = sizeNames.replace(changeElem['target'], changeElem['changeName'])
    # ---------------------------------------------------------------
    print('sizeNames:', sizeNames)

    mixtureNameList = []
    mixtureList = result['newMixtureRate']
    # print('mixtureList:',mixtureList)
    for mixtureElem in mixtureList:
        mixture = mixtureElem['material']
        mixtureNameList.append(mixture)
    mixtureNames = "|".join(mixtureNameList)

    # ----------------------------치환식------------------------
    for changeElem in changeList:
        mixtureNames = mixtureNames.replace(changeElem['target'], changeElem['changeName'])
    # ---------------------------------------------------------------
    print('mixtureNames:', mixtureNames)
    country = result['madeInCountry']
    print("country:", country)
    imageUrlList = []
    imageNameList = []
    images = result['goodsImages']
    count = 0
    timeNow = timeCreate.strftime("%Y%m%d")
    for image in images:
        imageUrl = image['imageUrl'].replace("https://image-cache.sinsang.market/images/",
                                             "https://image-v4.sinsang.market/?f=https://image-cache.sinsang.market/images/") + "&w=1500&h=2000"
        # print(imageUrl)

        countTwoDigit = ConvertToTwoDigits(count)
        productCountTwoDigit = ConvertToTwoDigits(productCount)
        imageName = timeNow + "_" + str(shopId) + "_" + str(productCountTwoDigit) + "_" + str(countTwoDigit) + ".jpg"
        imageNameList.append(imageName)

        timeToday = timeCreate.strftime("%Y%m%d")
        imageUrlList.append([imageUrl, "{}\\000_picture\\{}".format(timeToday, imageName)])
        count += 1
    imageNames = "|".join(imageNameList)
    print('imageNames:', imageNames)
    # pprint.pprint(imageUrlList)
    registerDate = result['regiDate'].replace("registered", "").replace(".", "").strip()
    print('registerDate:', registerDate)
    print("==================================")
    return registerDate, timeToday, name, price, category, colorNames, sizeNames, mixtureNames, country, imageNames, str(
        imageUrlList)


def createFolder(directory):
    try:
        if not os.path.exists(directory):
            os.makedirs(directory)
    except OSError:
        print('Error: Creating directory. ' + directory)


def GetCategory(cdIdx):
    categorys = [
        {
            "cdIdx": 1,
            "detailName": "Outer",
            "cgIdx": 1,
            "genderName": "Women's",
            "ciIdx": 1,
            "itemName": "Clothing"
        },
        {
            "cdIdx": 2,
            "detailName": "T-shirts&Tops",
            "cgIdx": 1,
            "genderName": "Women's",
            "ciIdx": 1,
            "itemName": "Clothing"
        },
        {
            "cdIdx": 3,
            "detailName": "Dresses",
            "cgIdx": 1,
            "genderName": "Women's",
            "ciIdx": 1,
            "itemName": "Clothing"
        },
        {
            "cdIdx": 4,
            "detailName": "Blouses",
            "cgIdx": 1,
            "genderName": "Women's",
            "ciIdx": 1,
            "itemName": "Clothing"
        },
        {
            "cdIdx": 5,
            "detailName": "Knitwear",
            "cgIdx": 1,
            "genderName": "Women's",
            "ciIdx": 1,
            "itemName": "Clothing"
        },
        {
            "cdIdx": 6,
            "detailName": "Jeans",
            "cgIdx": 1,
            "genderName": "Women's",
            "ciIdx": 1,
            "itemName": "Clothing"
        },
        {
            "cdIdx": 7,
            "detailName": "Pants",
            "cgIdx": 1,
            "genderName": "Women's",
            "ciIdx": 1,
            "itemName": "Clothing"
        },
        {
            "cdIdx": 8,
            "detailName": "Skirts",
            "cgIdx": 1,
            "genderName": "Women's",
            "ciIdx": 1,
            "itemName": "Clothing"
        },
        {
            "cdIdx": 9,
            "detailName": "Sales",
            "cgIdx": 1,
            "genderName": "Women's",
            "ciIdx": 1,
            "itemName": "Clothing"
        },
        {
            "cdIdx": 10,
            "detailName": "In Season",
            "cgIdx": 1,
            "genderName": "Women's",
            "ciIdx": 1,
            "itemName": "Clothing"
        },
        {
            "cdIdx": 11,
            "detailName": "Outer",
            "cgIdx": 2,
            "genderName": "Men's",
            "ciIdx": 1,
            "itemName": "Clothing"
        },
        {
            "cdIdx": 12,
            "detailName": "T-shirts&Tops",
            "cgIdx": 2,
            "genderName": "Men's",
            "ciIdx": 1,
            "itemName": "Clothing"
        },
        {
            "cdIdx": 13,
            "detailName": "Shirts",
            "cgIdx": 2,
            "genderName": "Men's",
            "ciIdx": 1,
            "itemName": "Clothing"
        },
        {
            "cdIdx": 14,
            "detailName": "Knitwear",
            "cgIdx": 2,
            "genderName": "Men's",
            "ciIdx": 1,
            "itemName": "Clothing"
        },
        {
            "cdIdx": 15,
            "detailName": "Suit",
            "cgIdx": 2,
            "genderName": "Men's",
            "ciIdx": 1,
            "itemName": "Clothing"
        },
        {
            "cdIdx": 16,
            "detailName": "Jeans",
            "cgIdx": 2,
            "genderName": "Men's",
            "ciIdx": 1,
            "itemName": "Clothing"
        },
        {
            "cdIdx": 17,
            "detailName": "Pants",
            "cgIdx": 2,
            "genderName": "Men's",
            "ciIdx": 1,
            "itemName": "Clothing"
        },
        {
            "cdIdx": 18,
            "detailName": "Sales",
            "cgIdx": 2,
            "genderName": "Men's",
            "ciIdx": 1,
            "itemName": "Clothing"
        },
        {
            "cdIdx": 19,
            "detailName": "In Season",
            "cgIdx": 2,
            "genderName": "Men's",
            "ciIdx": 1,
            "itemName": "Clothing"
        },
        {
            "cdIdx": 20,
            "detailName": "Flat Shoes/Loafers",
            "cgIdx": 1,
            "genderName": "Women's",
            "ciIdx": 2,
            "itemName": "Shoes"
        },
        {
            "cdIdx": 21,
            "detailName": "Heels/Pumps",
            "cgIdx": 1,
            "genderName": "Women's",
            "ciIdx": 2,
            "itemName": "Shoes"
        },
        {
            "cdIdx": 22,
            "detailName": "Wedge Heels",
            "cgIdx": 1,
            "genderName": "Women's",
            "ciIdx": 2,
            "itemName": "Shoes"
        },
        {
            "cdIdx": 23,
            "detailName": "Sandals/Slippers/Flip-flops",
            "cgIdx": 1,
            "genderName": "Women's",
            "ciIdx": 2,
            "itemName": "Shoes"
        },
        {
            "cdIdx": 24,
            "detailName": "Sneakers/Running Shoes",
            "cgIdx": 1,
            "genderName": "Women's",
            "ciIdx": 2,
            "itemName": "Shoes"
        },
        {
            "cdIdx": 25,
            "detailName": "Boots",
            "cgIdx": 1,
            "genderName": "Women's",
            "ciIdx": 2,
            "itemName": "Shoes"
        },
        {
            "cdIdx": 26,
            "detailName": "Handmade Shoes",
            "cgIdx": 1,
            "genderName": "Women's",
            "ciIdx": 2,
            "itemName": "Shoes"
        },
        {
            "cdIdx": 27,
            "detailName": "Sales",
            "cgIdx": 1,
            "genderName": "Women's",
            "ciIdx": 2,
            "itemName": "Shoes"
        },
        {
            "cdIdx": 28,
            "detailName": "Loafers/Flats",
            "cgIdx": 2,
            "genderName": "Men's",
            "ciIdx": 2,
            "itemName": "Shoes"
        },
        {
            "cdIdx": 29,
            "detailName": "Sneakers/Running Shoes",
            "cgIdx": 2,
            "genderName": "Men's",
            "ciIdx": 2,
            "itemName": "Shoes"
        },
        {
            "cdIdx": 30,
            "detailName": "Sandals/Slippers/Flip-flops",
            "cgIdx": 2,
            "genderName": "Men's",
            "ciIdx": 2,
            "itemName": "Shoes"
        },
        {
            "cdIdx": 31,
            "detailName": "Dress Shoes",
            "cgIdx": 2,
            "genderName": "Men's",
            "ciIdx": 2,
            "itemName": "Shoes"
        },
        {
            "cdIdx": 32,
            "detailName": "Handmade Shoes",
            "cgIdx": 2,
            "genderName": "Men's",
            "ciIdx": 2,
            "itemName": "Shoes"
        },
        {
            "cdIdx": 33,
            "detailName": "Boots",
            "cgIdx": 2,
            "genderName": "Men's",
            "ciIdx": 2,
            "itemName": "Shoes"
        },
        {
            "cdIdx": 34,
            "detailName": "Sales",
            "cgIdx": 2,
            "genderName": "Men's",
            "ciIdx": 2,
            "itemName": "Shoes"
        },
        {
            "cdIdx": 35,
            "detailName": "Children's Shoes/Shoes for Baby Walkers",
            "cgIdx": 3,
            "genderName": "Children's",
            "ciIdx": 2,
            "itemName": "Shoes"
        },
        {
            "cdIdx": 36,
            "detailName": "Girl's Shoes",
            "cgIdx": 3,
            "genderName": "Children's",
            "ciIdx": 2,
            "itemName": "Shoes"
        },
        {
            "cdIdx": 37,
            "detailName": "Boy's Shoes",
            "cgIdx": 3,
            "genderName": "Children's",
            "ciIdx": 2,
            "itemName": "Shoes"
        },
        {
            "cdIdx": 38,
            "detailName": "Sneakers/Running Shoes",
            "cgIdx": 3,
            "genderName": "Children's",
            "ciIdx": 2,
            "itemName": "Shoes"
        },
        {
            "cdIdx": 39,
            "detailName": "Flat Shoes/Loafers",
            "cgIdx": 3,
            "genderName": "Children's",
            "ciIdx": 2,
            "itemName": "Shoes"
        },
        {
            "cdIdx": 40,
            "detailName": "Wool/Boots",
            "cgIdx": 3,
            "genderName": "Children's",
            "ciIdx": 2,
            "itemName": "Shoes"
        },
        {
            "cdIdx": 41,
            "detailName": "Slippers/Overshoes",
            "cgIdx": 3,
            "genderName": "Children's",
            "ciIdx": 2,
            "itemName": "Shoes"
        },
        {
            "cdIdx": 42,
            "detailName": "Slippers/Sandals/Flip-flops",
            "cgIdx": 3,
            "genderName": "Children's",
            "ciIdx": 2,
            "itemName": "Shoes"
        },
        {
            "cdIdx": 43,
            "detailName": "Bags",
            "cgIdx": 3,
            "genderName": "Children's",
            "ciIdx": 3,
            "itemName": "Bags"
        },
        {
            "cdIdx": 44,
            "detailName": "Others",
            "cgIdx": 3,
            "genderName": "Children's",
            "ciIdx": 2,
            "itemName": "Shoes"
        },
        {
            "cdIdx": 45,
            "detailName": "Leather",
            "cgIdx": 1,
            "genderName": "Women's",
            "ciIdx": 3,
            "itemName": "Bags"
        },
        {
            "cdIdx": 46,
            "detailName": "Shoulder Bags",
            "cgIdx": 1,
            "genderName": "Women's",
            "ciIdx": 3,
            "itemName": "Bags"
        },
        {
            "cdIdx": 47,
            "detailName": "Tote Bags",
            "cgIdx": 1,
            "genderName": "Women's",
            "ciIdx": 3,
            "itemName": "Bags"
        },
        {
            "cdIdx": 48,
            "detailName": "Cross-body Bags",
            "cgIdx": 1,
            "genderName": "Women's",
            "ciIdx": 3,
            "itemName": "Bags"
        },
        {
            "cdIdx": 49,
            "detailName": "Clutch Bags/Wallets",
            "cgIdx": 1,
            "genderName": "Women's",
            "ciIdx": 3,
            "itemName": "Bags"
        },
        {
            "cdIdx": 50,
            "detailName": "Backpacks",
            "cgIdx": 1,
            "genderName": "Women's",
            "ciIdx": 3,
            "itemName": "Bags"
        },
        {
            "cdIdx": 51,
            "detailName": "Others",
            "cgIdx": 1,
            "genderName": "Women's",
            "ciIdx": 3,
            "itemName": "Bags"
        },
        {
            "cdIdx": 52,
            "detailName": "Sales",
            "cgIdx": 1,
            "genderName": "Women's",
            "ciIdx": 3,
            "itemName": "Bags"
        },
        {
            "cdIdx": 53,
            "detailName": "Shoulder Bags/Tote Bags",
            "cgIdx": 2,
            "genderName": "Men's",
            "ciIdx": 3,
            "itemName": "Bags"
        },
        {
            "cdIdx": 54,
            "detailName": "Cross-body Bags/Messenger Bags",
            "cgIdx": 2,
            "genderName": "Men's",
            "ciIdx": 3,
            "itemName": "Bags"
        },
        {
            "cdIdx": 55,
            "detailName": "Backpacks",
            "cgIdx": 2,
            "genderName": "Men's",
            "ciIdx": 3,
            "itemName": "Bags"
        },
        {
            "cdIdx": 56,
            "detailName": "Chalk Bags/Pouches",
            "cgIdx": 2,
            "genderName": "Men's",
            "ciIdx": 3,
            "itemName": "Bags"
        },
        {
            "cdIdx": 57,
            "detailName": "Wallets",
            "cgIdx": 2,
            "genderName": "Men's",
            "ciIdx": 3,
            "itemName": "Bags"
        },
        {
            "cdIdx": 58,
            "detailName": "Business",
            "cgIdx": 2,
            "genderName": "Men's",
            "ciIdx": 3,
            "itemName": "Bags"
        },
        {
            "cdIdx": 59,
            "detailName": "Others",
            "cgIdx": 2,
            "genderName": "Men's",
            "ciIdx": 3,
            "itemName": "Bags"
        },
        {
            "cdIdx": 60,
            "detailName": "Sales",
            "cgIdx": 2,
            "genderName": "Men's",
            "ciIdx": 3,
            "itemName": "Bags"
        },
        {
            "cdIdx": 61,
            "detailName": "Jewelry",
            "cgIdx": 1,
            "genderName": "Women's",
            "ciIdx": 4,
            "itemName": "Accessories"
        },
        {
            "cdIdx": 62,
            "detailName": "Belts",
            "cgIdx": 1,
            "genderName": "Women's",
            "ciIdx": 4,
            "itemName": "Accessories"
        },
        {
            "cdIdx": 63,
            "detailName": "Hairpin/band",
            "cgIdx": 1,
            "genderName": "Women's",
            "ciIdx": 4,
            "itemName": "Accessories"
        },
        {
            "cdIdx": 64,
            "detailName": "Hats",
            "cgIdx": 1,
            "genderName": "Women's",
            "ciIdx": 4,
            "itemName": "Accessories"
        },
        {
            "cdIdx": 65,
            "detailName": "Sunglasses/Glasses",
            "cgIdx": 1,
            "genderName": "Women's",
            "ciIdx": 4,
            "itemName": "Accessories"
        },
        {
            "cdIdx": 66,
            "detailName": "Watches",
            "cgIdx": 1,
            "genderName": "Women's",
            "ciIdx": 4,
            "itemName": "Accessories"
        },
        {
            "cdIdx": 67,
            "detailName": "Others",
            "cgIdx": 1,
            "genderName": "Women's",
            "ciIdx": 4,
            "itemName": "Accessories"
        },
        {
            "cdIdx": 68,
            "detailName": "Jewelry",
            "cgIdx": 2,
            "genderName": "Men's",
            "ciIdx": 4,
            "itemName": "Accessories"
        },
        {
            "cdIdx": 69,
            "detailName": "Hats",
            "cgIdx": 2,
            "genderName": "Men's",
            "ciIdx": 4,
            "itemName": "Accessories"
        },
        {
            "cdIdx": 70,
            "detailName": "Sunglasses",
            "cgIdx": 2,
            "genderName": "Men's",
            "ciIdx": 4,
            "itemName": "Accessories"
        },
        {
            "cdIdx": 71,
            "detailName": "Glasses",
            "cgIdx": 2,
            "genderName": "Men's",
            "ciIdx": 4,
            "itemName": "Accessories"
        },
        {
            "cdIdx": 72,
            "detailName": "Belts",
            "cgIdx": 2,
            "genderName": "Men's",
            "ciIdx": 4,
            "itemName": "Accessories"
        },
        {
            "cdIdx": 73,
            "detailName": "Watches",
            "cgIdx": 2,
            "genderName": "Men's",
            "ciIdx": 4,
            "itemName": "Accessories"
        },
        {
            "cdIdx": 74,
            "detailName": "Others",
            "cgIdx": 2,
            "genderName": "Men's",
            "ciIdx": 4,
            "itemName": "Accessories"
        },
        {
            "cdIdx": 75,
            "detailName": "Outer",
            "cgIdx": 4,
            "genderName": "Children's",
            "ciIdx": 1,
            "itemName": "Clothing"
        },
        {
            "cdIdx": 76,
            "detailName": "Knitwear",
            "cgIdx": 4,
            "genderName": "Children's",
            "ciIdx": 1,
            "itemName": "Clothing"
        },
        {
            "cdIdx": 77,
            "detailName": "T-shirts/Tops",
            "cgIdx": 4,
            "genderName": "Children's",
            "ciIdx": 1,
            "itemName": "Clothing"
        },
        {
            "cdIdx": 78,
            "detailName": "Jeans",
            "cgIdx": 4,
            "genderName": "Children's",
            "ciIdx": 1,
            "itemName": "Clothing"
        },
        {
            "cdIdx": 79,
            "detailName": "Dresses",
            "cgIdx": 4,
            "genderName": "Children's",
            "ciIdx": 1,
            "itemName": "Clothing"
        },
        {
            "cdIdx": 80,
            "detailName": "Pants",
            "cgIdx": 4,
            "genderName": "Children's",
            "ciIdx": 1,
            "itemName": "Clothing"
        },
        {
            "cdIdx": 81,
            "detailName": "Blouses",
            "cgIdx": 4,
            "genderName": "Children's",
            "ciIdx": 1,
            "itemName": "Clothing"
        },
        {
            "cdIdx": 82,
            "detailName": "Skirts",
            "cgIdx": 4,
            "genderName": "Children's",
            "ciIdx": 1,
            "itemName": "Clothing"
        },
        {
            "cdIdx": 83,
            "detailName": "Sales",
            "cgIdx": 4,
            "genderName": "Children's",
            "ciIdx": 1,
            "itemName": "Clothing"
        },
        {
            "cdIdx": 84,
            "detailName": "In Season",
            "cgIdx": 4,
            "genderName": "Children's",
            "ciIdx": 1,
            "itemName": "Clothing"
        },
        {
            "cdIdx": 85,
            "detailName": "Outer",
            "cgIdx": 5,
            "genderName": "Children's",
            "ciIdx": 1,
            "itemName": "Clothing"
        },
        {
            "cdIdx": 86,
            "detailName": "Shirts",
            "cgIdx": 5,
            "genderName": "Children's",
            "ciIdx": 1,
            "itemName": "Clothing"
        },
        {
            "cdIdx": 87,
            "detailName": "T-shirts/Tops",
            "cgIdx": 5,
            "genderName": "Children's",
            "ciIdx": 1,
            "itemName": "Clothing"
        },
        {
            "cdIdx": 88,
            "detailName": "Jeans",
            "cgIdx": 5,
            "genderName": "Children's",
            "ciIdx": 1,
            "itemName": "Clothing"
        },
        {
            "cdIdx": 89,
            "detailName": "Knitwear",
            "cgIdx": 5,
            "genderName": "Children's",
            "ciIdx": 1,
            "itemName": "Clothing"
        },
        {
            "cdIdx": 90,
            "detailName": "Pants",
            "cgIdx": 5,
            "genderName": "Children's",
            "ciIdx": 1,
            "itemName": "Clothing"
        },
        {
            "cdIdx": 91,
            "detailName": "Suit",
            "cgIdx": 4,
            "genderName": "Children's",
            "ciIdx": 1,
            "itemName": "Clothing"
        },
        {
            "cdIdx": 92,
            "detailName": "In Season",
            "cgIdx": 5,
            "genderName": "Children's",
            "ciIdx": 1,
            "itemName": "Clothing"
        },
        {
            "cdIdx": 93,
            "detailName": "Sales",
            "cgIdx": 5,
            "genderName": "Children's",
            "ciIdx": 1,
            "itemName": "Clothing"
        },
        {
            "cdIdx": 94,
            "detailName": "Set Products",
            "cgIdx": 1,
            "genderName": "Women's",
            "ciIdx": 1,
            "itemName": "Clothing"
        },
        {
            "cdIdx": 95,
            "detailName": "Shirts",
            "cgIdx": 1,
            "genderName": "Women's",
            "ciIdx": 1,
            "itemName": "Clothing"
        },
        {
            "cdIdx": 96,
            "detailName": "Plus Size",
            "cgIdx": 1,
            "genderName": "Women's",
            "ciIdx": 1,
            "itemName": "Clothing"
        },
        {
            "cdIdx": 97,
            "detailName": "Maternity Clothing",
            "cgIdx": 1,
            "genderName": "Women's",
            "ciIdx": 1,
            "itemName": "Clothing"
        },
        {
            "cdIdx": 98,
            "detailName": "Top&Bottom",
            "cgIdx": 4,
            "genderName": "Children's",
            "ciIdx": 1,
            "itemName": "Clothing"
        }
    ]
    categoryResult = ""
    for category in categorys:
        if str(category['cdIdx']) == str(cdIdx):
            categoryResult = category['detailName']
            print(categoryResult)
            break
    return categoryResult


def load_excel_storeinfo(fname):
    fname = 'storelist.xlsx'
    wb = openpyxl.load_workbook(fname)
    ws = wb.active
    no_row = ws.max_row
    print("행갯수:", no_row)
    data_list = []
    for i in range(2, no_row + 1):
        url = ws.cell(row=i, column=1).value
        if url == "" or url == None:
            print('데이타 더 이상 없음')
            break
        size = ws.cell(row=i, column=2).value
        storeName = ws.cell(row=i, column=3).value
        storeAddress = ws.cell(row=i, column=4).value
        storeStyle = ws.cell(row=i, column=5).value
        regex = re.compile("\d+")
        storeNo = regex.findall(url)[0]
        data = {'url': url, 'size': size, 'storeName': storeName, 'storeAddress': storeAddress,
                'storeStyle': storeStyle, 'storeNo': storeNo}
        data_list.append(data)
    print(data_list)
    return data_list


def CategoryChangeList(fname):
    wb = openpyxl.load_workbook(fname)
    ws = wb.active
    no_row = ws.max_row
    print("행갯수:", no_row)
    data_list = []
    for i in range(2, no_row + 1):
        target = ws.cell(row=i, column=1).value
        if target == "" or target == None:
            print('데이타 더 이상 없음')
            break
        changeName = ws.cell(row=i, column=2).value
        data = {'target': target, 'changeName': changeName}
        data_list.append(data)
    print("changeList:", data_list)
    return data_list


def load_excel_downloadinfo(fname):
    print("다운22", fname)
    wb = openpyxl.load_workbook(fname)
    ws = wb.active
    no_row = ws.max_row
    print("행갯수:", no_row)
    data_list = []
    for i in range(2, no_row + 1):
        downInfos = eval(ws.cell(row=i, column=15).value)
        # print(downInfos)
        if downInfos == "" or downInfos == None:
            print('데이타 더 이상 없음')
            break
        for downInfo in downInfos:
            downUrl = downInfo[0]
            saveName = downInfo[1]
            data = {'downUrl': downUrl, 'saveName': saveName}
            data_list.append(data)

    print('downInfos:', data_list)
    inputString = data_list[0]['saveName']
    substring = inputString.split("\\000_picture")[0] + "\\000_picture"
    print(substring)
    createFolder(substring)

    print("그림갯수:", len(data_list))

    return data_list


def getAllFilenames(folder_path):
    filenames = []
    for filename in os.listdir(folder_path):
        if os.path.isfile(os.path.join(folder_path, filename)):
            filenames.append(filename)
    return filenames


def GetFileNames(folder_path):
    file_names = []
    for dirpath, _, filenames in os.walk(folder_path):
        for filename in filenames:
            file_path = os.path.join(dirpath, filename)
            file_names.append(file_path)
    return file_names


# ---------------------------주요 FUNCTION
def GetExcelResult(self, id, pw, sleepingTime, speedMin, continueFlag, speedMax, sleepingCount):
    # 샵정보가져오기
    storeInfoFile = 'storelist.xlsx'
    storeInfos = load_excel_storeinfo(storeInfoFile)
    print("storeInfos:", storeInfos)
    accessToken = GetToken(id, pw)
    print("accessToken:", accessToken)
    print("continueFlag:", continueFlag, "/ continueFlag_TYPE:", type(continueFlag))
    if continueFlag == True:  # 이어하기 Flag가 잇으면 해당 상품 전부터 시작
        with open('lastOne.json', "r", encoding='utf-8-sig') as f:
            lastOne = json.load(f)
        index = None
        for i, item in enumerate(storeInfos):
            if item['storeName'] == lastOne['name']:
                index = i - 1
                break
        if index == -1:  # 만약 첨부터 시작했다면 그냥 첨부터 또 함
            index = 0
        storeInfos = storeInfos[index:]
        print('changedStoreInfos:', storeInfos)

    wb = openpyxl.Workbook()
    ws = wb.active
    columnName = ['상품등록일', '날짜', '매장이름', '매장명', '호수', '스타일', '상품명', '가격', '카테고리', '색상', '사이즈', '혼용률', '제조국가', '이미지',
                  '이미지URL']
    ws.append(columnName)
    runningCount = 0
    timeCreate = datetime.datetime.now()
    timeCreateDay = timeCreate.strftime("%Y%m%d")
    createFolder('{}\{}'.format(timeCreateDay, "000_picture"))
    for storeInfo in storeInfos:
        shopId = storeInfo['storeNo']
        size = storeInfo['size']

        try:
            productIdList, storeNameWeb = GetProuctList(accessToken, shopId, size)
        except:
            continue
        text = "스토어명 : {} / 크롤링 중...".format(storeNameWeb)
        self.user_signal.emit(text)
        lastStore = {'name': storeInfo['storeName']}
        with open('lastOne.json', 'w', encoding='utf-8-sig') as f:
            json.dump(lastStore, f, indent=2, ensure_ascii=False)
        print('productIdList:', productIdList)
        for productCount, productId in enumerate(productIdList):
            url = 'https://sinsangmarket.kr/store/{}?sort=DATE&isPublic=true&modalGid={}'.format(storeInfo['storeNo'],
                                                                                                 productId)
            print('productUrl:', url)

            try:
                registerDate, timeToday, name, price, category, colorNames, sizeNames, mixtureNames, country, imageNames, imageUrlList = GetDetail(
                    accessToken, productId, productCount, shopId, timeCreate)
                time.sleep(random.randint(speedMin, speedMax))
            except:
                print("에러")
                continue
            print(storeInfo)

            data = [registerDate, timeCreateDay, storeNameWeb, storeInfo['storeName'], storeInfo['storeAddress'],
                    storeInfo['storeStyle'], name, price, category, colorNames, sizeNames, mixtureNames, country,
                    imageNames, imageUrlList]
            ws.append(data)

            wb.save('{}\\{}_전체상품리스트.xlsx'.format(timeCreateDay, timeToday))
            print("===================={}===================".format(storeInfo['storeName']))
        runningCount += 1

        if runningCount >= 1 and runningCount % sleepingCount == 0:
            print("슬립중...{}".format(runningCount))
            time.sleep(sleepingTime * 60)
            accessToken = GetToken(id, pw)
            print("accessToken:", accessToken)

    text = "작업완료"
    self.user_signal.emit(text)


def GetDownload(self, fnameDownload):
    headers = {
        'User-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/114.0.0.0 Safari/537.36', }

    downInfos = load_excel_downloadinfo(fnameDownload)
    savePath = downInfos[0]['saveName'].split("\\")[0] + "\\" + downInfos[0]['saveName'].split("\\")[1]

    fileNames = GetFileNames(savePath)
    with open('fileNames.json', 'w', encoding='utf-8-sig') as f:
        json.dump(fileNames, f, indent=2, ensure_ascii=False)
    newDownInfos = []
    for downInfo in downInfos:
        if downInfo['saveName'] not in fileNames:
            newDownInfos.append(downInfo)
    downInfos = newDownInfos

    wb = openpyxl.Workbook()
    ws = wb.active

    for downInfo in downInfos:
        print(downInfo['saveName'])
        ws.append([downInfo['saveName']])
    wb.save('test.xlsx')
    print("갯수는:", len(downInfos))

    print('downInfos:', downInfos)

    with open('downInfos.json', 'w', encoding='utf-8-sig') as f:
        json.dump(downInfos, f, indent=2, ensure_ascii=False)
    for index, downInfo in enumerate(downInfos):
        text = f"{index + 1}/{len(downInfos)}다운로드 중..."
        print(text)
        self.user_signal.emit(text)
        print("imageUrl", downInfo['downUrl'])
        try:
            image_res = requests.get(downInfo['downUrl'], headers=headers)  # 그림파일 저장
            image_res.raise_for_status()
        except:
            text = "다운로드실패"
            print(text)
            self.user_signal.emit(text)
            continue

        with open(downInfo['saveName'], "wb") as f:
            f.write(image_res.content)  # 그림파일 각각 저장
        speed = 1

        text = "다운로드완료"
        print(text)
        self.user_signal.emit(text)

    print("작업완료")


def RemoveBef(self, fnameDouble):
    # 기존정보가져오기
    folderName = 'reference'
    fileNames = getAllFilenames(folderName)
    # print('fileNames:',fileNames)
    df2 = pd.DataFrame()
    for fileName in fileNames:
        print('fileName:', fileName)
        data = pd.read_excel(folderName + "\\" + fileName)
        try:
            data['상품등록일'] = data['상품등록일'].str.replace('.', '')
        except:
            print("점 없음")
        try:
            if 'Num' in data.columns:
                data.drop('Num', axis=1, inplace=True)
        except:
            print("Num없음")
        try:
            if '이미지URL' in data.columns:
                index_of_image_url = data.columns.get_loc('이미지URL')
                data = data.iloc[:, :index_of_image_url + 1]
        except:
            print("뒤에더없음")

        print(data)
        df2 = pd.concat([df2, data])
        print("합치기")
    print('df2:', df2)
    # 현재정보가져오기

    print("과거 정보 가져와서 합치기")
    # timeToday=datetime.datetime.now().strftime("%Y%m%d")
    # folderName = timeToday
    # fileNames = getAllFilenames(folderName)
    # for fileName in fileNames:
    #     df1=pd.read_excel(folderName+"\\"+fileName)
    df1 = pd.read_excel(fnameDouble)

    print("df1:", df1)
    # try:
    #     df1['상품등록일'] = df1['상품등록일'].str.replace('.', '')
    # except:
    #     print("점없음2")
    merged_df = df1.merge(df2, on=['상품명', '매장명', '가격'], how='left', indicator=True, suffixes=("", "_duplicate"))
    df1 = merged_df[merged_df['_merge'] == 'left_only']
    df1 = df1.drop(columns='_merge')
    df1 = df1.loc[:, ~df1.columns.str.endswith('_duplicate')]
    df1.to_excel(fnameDouble, index=False)
    text = "중복제거완료"
    self.user_signal.emit(text)


# -----------------------------주요 실행

def GetLogOut(loginId):
    headers = {
        'authority': 'iteratehq.com',
        'accept': '*/*',
        'accept-language': 'ko-KR,ko;q=0.9,en-US;q=0.8,en;q=0.7',
        'authorization': 'Bearer eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJjb21wYW55X2lkIjoiNjI3OGI3ZDUyY2Q2YmYwMDAxZDM4NDkyIiwidXNlcl9pZCI6IjY0MDA0YmVmMmNhM2I2MDAwMTRmN2I0ZSIsInVzZXJfaWRfZXh0Ijp0cnVlLCJpYXQiOjE3MDAyMjIyMjB9.yj_JVTLvsnt1KufvWIxAAT4PWL6yni36kC_94O6rObA',
        'content-type': 'application/json',
        'origin': 'https://sinsangmarket.kr',
        'referer': 'https://sinsangmarket.kr/',
        'sec-ch-ua': '"Google Chrome";v="119", "Chromium";v="119", "Not?A_Brand";v="24"',
        'sec-ch-ua-mobile': '?0',
        'sec-ch-ua-platform': '"Windows"',
        'sec-fetch-dest': 'empty',
        'sec-fetch-mode': 'cors',
        'sec-fetch-site': 'cross-site',
        'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/119.0.0.0 Safari/537.36',
    }

    json_data = {
        'browser': {
            'url': 'https://sinsangmarket.kr/login',
        },
        'user_traits': {
            'external_id': loginId,
            'store_type': 'SOME',
            'store_id': '415101',
            'job_class': 'SCEO',
            'platform': 'WEB',
            'lang': 'ko',
        },
        'tracking': {
            'last_updated': 1697808949,
        },
        'app': {
            'ui_style': 'web',
        },
        'targeting': {},
        'type': 'web',
        'v': 20160516,
    }

    response = requests.post('https://iteratehq.com/api/v1/surveys/embed', headers=headers, json=json_data)
    print(response.text)

def GetGoogleSpread():
    scope = 'https://spreadsheets.google.com/feeds'
    json = 'credential.json'
    credentialdict={
  "type": "service_account",
  "project_id": "autosinsang",
  "private_key_id": "29840e08f2c51113f5eec4a60308b512396eac8a",
  "private_key": "-----BEGIN PRIVATE KEY-----\nMIIEvQIBADANBgkqhkiG9w0BAQEFAASCBKcwggSjAgEAAoIBAQDH90Api5+29/Tj\n0jo3QNr7V/6xEKN4Q+2iPX++Yy/osJT2HuiEeQ7KVOohsI0GeSvOo8FUzJBZQviB\nfTrH1bY+EKWcVW52RlEXtBXSzTrhMQMjcf2VjwABqytah3bHBjrXu1dcsPylCtUL\n/EqkjFApppZieKcm/R6N5Wa3z2f71yr5Si44UIM8I4sz7kAm52+muw1r/YTSjWn7\ng3DerbGekImF0lbQB7fkKDzP2MmU4ja3NGWOZSN1+wEp10IN7dzRkYzT9ErRSN2O\nuFIB9mIpAv0dBZsc1i94USVajV08pnxW6hc2+7Kh0bEe9VLZOOywfJCHhqaoxtHq\nrd8Im6AxAgMBAAECggEAANcP7STBlvPdPgKijgziqmWN8PrGAs4IyD3znxexoOJc\nJjPhQpegseeMJIdINtzfcYOsxJnD6Pu9I9c+pYLhiWycpDXMd2xuMC7A71XZ0Xtz\ngMjYv1WkLUUlr0vZeGO+kWjDRuUGIhwqbgrDHdcqZDKFpzQfGqcWCTHLp3Nu0LgO\nfgV5SPXPVRfPN9wvHjRfq6MZqLc79D7+RJuCCtdOWGI8a842jLlu12T+AICOkDm6\nVk6T8StNRE9khJUeLqzyZkNyeiJs0RuEKkVogjxkOMI3GL0w/YRZfCWbff0FZuIy\nGAYNxlHHZxYjb68NkItwnwCE6dBb+LG4hm5AxPp+CQKBgQDr6FahI0EUCl4TZTMT\nl4/Hrjj/uPerco7qrie3p+QvQDNThNhEdTMkuhdokL9ixmtG023Zun5Q59VdPZGQ\nRZU9Xi+5NNhMpzxPK/c8CkqfWU4N00C9Slv6B5isMWCOq5M/V0r4F5/ylz4/d9cW\nhJrFDFGC6zwoEx4nQORQWA2G6QKBgQDY/z+h0V0uwo6Pb0y75ifXG0RdN5HC6g4x\n5jPersg9/wMrJkhwR+GpI//D57nrwAV1RMPDu/AZDb/L99Kzf8xRnh4JMzG5XtqG\nVqBIbotfuyqcz9JeH8WDXIVsn0e/8xkLCYHCyaiQ2YaFn76VvudUwYK+fMX5Kiad\nBtl/wwmSCQKBgDVrDvK+AOaA0nyh1ccB8tTMZqvBPpPcBPCvLt8vbUQlzXmEA1S9\nFS+j09khrbS4KWKITb+XszKuzGmon2WD3B9hTH0rsi7pT4a99eXqjqs2la4bv/Uc\n+WbMgY28QJkN0PHu2lMDRnywMrA8er/mlwU3nN2AVsbcTV+mGgz3Y1jBAoGAZati\nMv0XPbtmdAnf2AHEFOLvqwVIRBzJvdNv5hvaMkHzSBpNwl3LeyYkxbMixzRW2lT3\nrBO4MOTiQjzA+d0c6/i7dfAIkiPBG0QnIQi6QIY/Nu2gUP1t21hUI0qVoXfS+frg\nU1WMpyFkB+OKtjqVxH5ONdXcJBv9r4DwEw5cw5ECgYEAmi5Rao9B4w1pC9qk1Z1q\nRGCCs1qBCzh61i8D9gDiZK63FQ8OJzs9ssVFpzmwgk/NxPPlxWN1Pf0LA4nTYKUT\nu45VhkbFof5kqes9mqyd5CZM5QnOkpNd2V7DNsP2Ff9uhdL3nLgTs8j/25L8RCX2\nONLluzkxTnLHB9J8H+sxD2g=\n-----END PRIVATE KEY-----\n",
  "client_email": "autosinsang@autosinsang.iam.gserviceaccount.com",
  "client_id": "107707675164618691222",
  "auth_uri": "https://accounts.google.com/o/oauth2/auth",
  "token_uri": "https://oauth2.googleapis.com/token",
  "auth_provider_x509_cert_url": "https://www.googleapis.com/oauth2/v1/certs",
  "client_x509_cert_url": "https://www.googleapis.com/robot/v1/metadata/x509/autosinsang%40autosinsang.iam.gserviceaccount.com",
  "universe_domain": "googleapis.com"
    }
    credentials = ServiceAccountCredentials.from_json_keyfile_dict(credentialdict, scope)
    gc = gspread.authorize(credentials)
    sheet_url = 'https://docs.google.com/spreadsheets/d/1TRwasMvoiihlGruPXxFQAuB_epKc1FzDm2t1uGt1PFc/edit#gid=0'
    doc = gc.open_by_url(sheet_url)
    worksheet = doc.worksheet('Sheet1')
    #=================전체정보가져오기
    all_data=worksheet.get_all_records()
    # pprint.pprint(all_data)
    # 현재 날짜와 요일 구하기
    now = datetime.datetime.now()
    days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
    current_day_name = days[now.weekday()]
    # 오늘 크롤링된 데이터만 필터링
    today_crawled_data = [item for item in all_data if item['크롤링 요일'] == current_day_name]

    # 변경할 키와 새 키를 매핑하는 딕셔너리
    key_mapping = {
        '매장주소': 'url',
        '상품 개수': 'size',
        '상호': 'storeName',
        '호수': 'storeAddress',
        '스타일': 'storeStyle',
        '크롤링 요일': 'ReqDay'
    }

    # 새 키를 사용하여 데이터 리스트를 업데이트
    today_crawled_data = [{key_mapping[k]: v for k, v in item.items()} for item in today_crawled_data]

    # 각 요소의 'url'에서 연속된 숫자를 추출하고 'storeNo' 키로 추가
    for item in today_crawled_data:
        # 정규식을 사용하여 URL에서 숫자를 추출
        numbers = re.findall(r'\d+', item['url'])
        if numbers:
            # 첫 번째 연속된 숫자를 'storeNo'로 설정
            item['storeNo'] = numbers[0]
        else:
            item['storeNo'] = 0

    print("today_crawled_data:",today_crawled_data,"/ today_crawled_data_TYPE:",type(today_crawled_data))
    #아이디비번가져오기
    worksheet = doc.worksheet('Sheet2')
    # =================전체정보가져오기
    all_data = worksheet.get_all_records()
    loginData=all_data
    print("loginData:",loginData,"/ loginData_TYPE:",type(loginData))
    return today_crawled_data,loginData

def SendMail(filepath,username,password,receiver):
    smtp_server = 'smtp.naver.com'
    smtp_port = 587

    # 네이버 이메일 계정 정보
    # username = 'mike102jiro@naver.com'  # 클라이언트 정보 입력
    # password = 'Jan240109$$$'  # 클라이언트 정보 입력

    # username = 'hellfir2@naver.com'  # 클라이언트 정보 입력
    # password = 'dlwndwo1!'  # 클라이언트 정보 입력
    # =================커스터마이징
    try:
        to_mail = receiver
    except:
        print("메일주소없음")
        return
    # =================
    # 메일 수신자 정보
    to_email = receiver
    # 참조자 정보
    cc_email = 'ljj3347@naver.com'

    # 메일 본문 및 제목 설정
    contentList=[]

    content="\n".join(contentList)


    # MIMEMultipart 객체 생성
    timeNow=datetime.datetime.now().strftime("%Y년%m월%d일 %H시%M분%S초")
    msg = MIMEMultipart('alternative')
    msg["Subject"] = "[결과](현재시각:{})".format(timeNow)  # 메일 제목
    msg['From'] = username
    msg['To'] = to_email
    msg['Cc'] = cc_email  # 참조 이메일 주소 추가
    msg.attach(MIMEText(content, 'plain'))

    # filepath='20240130_RESULT.xlsx'
    # 파일 첨부
    part = MIMEBase('application', 'octet-stream')
    with open(filepath, 'rb') as file:
        part.set_payload(file.read())
    encoders.encode_base64(part)
    part.add_header('Content-Disposition', f'attachment; filename={filepath}')
    msg.attach(part)

    # SMTP 서버 연결 및 로그인
    server = smtplib.SMTP(smtp_server, smtp_port)
    server.starttls()
    server.login(username, password)
    # 이메일 전송 (수신자와 참조자 모두에게 전송)
    to_and_cc_emails = [to_email] + [cc_email]
    server.sendmail(username, to_and_cc_emails, msg.as_string())
    # SMTP 서버 연결 종료
    server.quit()
    print("전송완료")


#==========작동부
filepath=""
username='vonschwarz90@naver.com'
password='S@aw7NG~wPV2B.5'
receiver='krrr@apmmust.com'
speedMin=5
speedMax=17
startTime=8
while True:
    currentTime=datetime.datetime.now().strftime("%H%M%S")
    print("현재시간:",currentTime)
    if currentTime=="200000":
        while True:
            try:
                storeInfos, loginData = GetGoogleSpread()
                break
            except:
                print("구글스프레드에러")
                time.sleep(1)

        print("storeInfos:", storeInfos, "/ storeInfos_TYPE:", type(storeInfos))
        accessToken = GetToken(loginData[0]['아이디'], loginData[0]['비밀번호'])
        print("accessToken:", accessToken, "/ accessToken_TYPE:", type(accessToken))

        wb = openpyxl.Workbook()
        ws = wb.active
        columnName = ['상품등록일', '날짜', '매장이름', '매장명', '호수', '스타일', '상품명', '가격', '카테고리', '색상', '사이즈', '혼용률', '제조국가', '이미지',
                      '이미지URL']
        ws.append(columnName)
        timeCreate = datetime.datetime.now()
        timeCreateDay = timeCreate.strftime("%Y%m%d")
        timeSave = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        createFolder('{}\{}'.format(timeCreateDay, "000_picture"))
        with open('storeInfos.json', 'w', encoding='utf-8-sig') as f:
            json.dump(storeInfos, f, indent=2, ensure_ascii=False)

        for storeInfo in storeInfos:
            shopId = storeInfo['storeNo']
            size = storeInfo['size']
            try:
                productIdList, storeNameWeb = GetProuctList(accessToken, shopId, size)
            except:
                continue
            text = "스토어명 : {} / 크롤링 중...".format(storeNameWeb)
            print(text)
            for productCount, productId in enumerate(productIdList):
                isTiming = is_time_between_8pm_and_5am()
                timeNow=datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
                print("작업여부:", isTiming,"현재시간:",timeNow)
                if isTiming==False: # 테스트 아닐때 풀어줘
                    print("작업시간아님1")
                    break
                timeNow = datetime.datetime.now()

                url = 'https://sinsangmarket.kr/store/{}?sort=DATE&isPublic=true&modalGid={}'.format(
                    storeInfo['storeNo'],
                    productId)
                print('productUrl:', url)

                try:
                    registerDate, timeToday, name, price, category, colorNames, sizeNames, mixtureNames, country, imageNames, imageUrlList = GetDetail(
                        accessToken, productId, productCount, shopId, timeCreate)
                    time.sleep(random.randint(speedMin, speedMax))
                except:
                    print("에러")
                    continue
                print(storeInfo)

                data = [registerDate, timeCreateDay, storeNameWeb, storeInfo['storeName'], storeInfo['storeAddress'],
                        storeInfo['storeStyle'], name, price, category, colorNames, sizeNames, mixtureNames, country,
                        imageNames, imageUrlList]
                ws.append(data)

                filepath = '{}_RESULT.xlsx'.format(timeSave)
                wb.save(filepath)
                print("===================={}===================".format(storeInfo['storeName']))
            if isTiming == False:
                print("작업시간아님2")
                break
        time.sleep(10)
        try:
            print("filepath:", filepath, "/ filepath_TYPE:", type(filepath))
            SendMail(filepath, username, password, receiver)
        except:
            print("메일전송실패")
        try:
            GetLogOut(loginData[0]['아이디'])
            print("로그아웃완료")
        except:
            print("로그아웃에러")

    else:
        print("대기중...")
    time.sleep(0.9)
